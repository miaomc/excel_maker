# coding=utf-8
import openpyxl
import copy


class Cell(object):
    def __init__(self):
        self._style = None
        self.value = None
        self.coordinate = None
        self.font = {'name': u'微软雅黑', 'sz': 9, 'bold': False, 'charset': 134,'color':None}
        self.border = {'left_style': None, 'right_style': None, 'top_style': None, 'bottom_style': None}
        self.alignment = {'horizontal': 'center', 'vertical': 'center', 'wraptext': True}  # 换行
        self.fill = {'color': None}
        self.number_format = u'常规'


def read_template(filename):
    workbook = openpyxl.load_workbook(filename)
    return workbook.active


def load_styles(sheet):
    """
    Font：设置字体大小、颜色、下划线等等
    Border：设置单元格的边框
    Alignment：单元格对齐
    img
    #PatternFill：设置图案或者颜色渐变
    # Protection：保护工作表
    """
    # new_book = openpyxl.Workbook()
    # new_sheet = new_book.active

    # 获取style,存入style_list[][] = Cell()的二维表
    style_list = []

    for row in range(1, sheet.max_row + 1):  # 1,2,...
        style_list.append([])
        for column in range(1, sheet.max_column + 1):  # A,B,... To be continue...
            style_list[row - 1].append(None)
            coordinate = chr(column + 64) + str(row)
            source_cell = sheet[coordinate]
            if sheet[coordinate].has_style:
                cell = Cell()
                cell.coordinate = source_cell.coordinate
                cell.value = source_cell.value
                cell._style = source_cell._style
                cell.font["name"] = source_cell.font.name
                cell.font["sz"] = source_cell.font.sz
                cell.font["bold"] = source_cell.font.bold
                if source_cell.font.color:
                    cell.font['color'] = source_cell.font.color.rgb
                cell.font["charset"] = source_cell.font.charset
                cell.border["left_style"] = source_cell.border.left.style
                cell.border["right_style"] = source_cell.border.right.style
                cell.border["top_style"] = source_cell.border.top.style
                cell.border["bottom_style"] = source_cell.border.bottom.style
                cell.fill["color"] = source_cell.fill.fgColor.rgb
                # cell.fill['tint'] =source_cell.fill.fgColor.tint
                cell.number_format = source_cell.number_format
                cell.alignment['horizontal'] = source_cell.alignment.horizontal
                cell.alignment['vertical'] = source_cell.alignment.vertical
                cell.alignment['wraptext'] = source_cell.alignment.wrapText

                style_list[row - 1][column - 1] = cell
    #             print([i.font['sz'] for i in style_list[0]])
    #             # print(row,column,style_list[row - 1][column - 1].coordinate, style_list[row - 1][column - 1].font['sz'])
    #     print(1,[i.font['sz'] for i in style_list[0]])
    #
    #     break
    # print(1,[i.font['sz'] for i in style_list[1]])

    # new_sheet[coordinate] = str(cell.fill["color"])

    # # tab颜色
    # new_sheet.sheet_properties.tabColor = sheet.sheet_properties.tabColor
    #
    # # 开始处理合并单元格形式为“(<CellRange A1：A4>,)，替换掉(<CellRange 和 >,)' 找到合并单元格
    # wm = list(sheet.merged_cells)
    # if len(wm) > 0:
    #     for i in range(0, len(wm)):
    #         cell2 = str(wm[i]).replace('(<CellRange ', '').replace('>,)', '')
    #         new_sheet.merge_cells(cell2)

    # new_book.save('0.xlsx')

    # for ii,i in enumerate(style_list):
    #     for jj,j in enumerate(i):
    #         print(ii,jj,j.font['sz'],end=' ')
    #     print()
    return style_list


def make_excel(sheet, style_list, filename):
    new_book = openpyxl.Workbook()
    new_sheet = new_book.active

    # tab颜色
    new_sheet.sheet_properties.tabColor = sheet.sheet_properties.tabColor

    # 调整高和宽
    for i, row in enumerate(sheet.iter_rows()):
        new_sheet.row_dimensions[i + 1].height = sheet.row_dimensions[i + 1].height
        for j, cell in enumerate(row):
            new_sheet.column_dimensions[chr(j + 65)].width = sheet.column_dimensions[chr(j + 65)].width
            new_sheet.cell(row=i + 1, column=j + 1, value=cell.value)



    for row in style_list:
        for cell in row:

            if not isinstance(new_sheet[cell.coordinate], openpyxl.cell.cell.MergedCell):
                # print(cell.coordinate, cell.font['sz'], cell.font['bold'], new_sheet[cell.coordinate],cell.fill['color'])
                new_sheet[cell.coordinate].value = cell.value
                new_sheet[cell.coordinate]._style = cell._style
                # print(cell.coordinate,cell.font['color'])
                try:
                    new_sheet[cell.coordinate].font = openpyxl.styles.Font(name=cell.font["name"], size=cell.font['sz'],
                                                                       bold=cell.font['bold'],color=cell.font['color'])
                except Exception as msg:
                    # print(cell.coordinate,msg)
                    pass
                # print(new_sheet[cell.coordinate].coordinate, new_sheet[cell.coordinate].font.size)
                new_sheet[cell.coordinate].border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(border_style=cell.border["left_style"]),
                    right=openpyxl.styles.Side(border_style=cell.border["right_style"]),
                    top=openpyxl.styles.Side(border_style=cell.border["top_style"]),
                    bottom=openpyxl.styles.Side(border_style=cell.border["bottom_style"]))
                # print(cell.coordinate, cell.border)
                new_sheet[cell.coordinate].fill = openpyxl.styles.PatternFill(patternType='solid',
                                                                              fgColor=cell.fill['color'])
                new_sheet[cell.coordinate].number_format = cell.number_format
                new_sheet[cell.coordinate].alignment = openpyxl.styles.Alignment(
                    horizontal=cell.alignment["horizontal"],
                    vertical=cell.alignment['vertical'],
                    wrap_text=cell.alignment['wraptext'])

    # 开始处理合并单元格形式为“(<CellRange A1：A4>,)，替换掉(<CellRange 和 >,)' 找到合并单元格
    wm = list(sheet.merged_cells)
    if len(wm) > 0:
        for i in range(0, len(wm)):
            cell2 = str(wm[i]).replace('(<CellRange ', '').replace('>,)', '')
            new_sheet.merge_cells(cell2)

    new_book.save(filename)


def main():
    worksheet = read_template('template.xlsx')
    style_list = load_styles(worksheet)
    make_excel(worksheet, style_list, '1.xlsx')
    return style_list


if __name__ == '__main__':
    l = main()
